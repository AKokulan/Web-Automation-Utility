import openpyxl,datetime,time
import pandas as pd

def load_workbook(workbook_path):
    workbook_path=workbook_path
    wb=""
    try:
        wb = openpyxl.load_workbook(workbook_path,read_only=False)
        return wb
    except:
        error="Unable to load excel workbook with the error: ",wb
        print(error)

def get_sheet_by_name(workbook_instance,sheet_name):
    wb=workbook_instance
    sheet_name=sheet_name
    try:
        sheet = wb.get_sheet_by_name(sheet_name)
        return sheet
    except:
        error="unable to open excel sheet"
        print(error)

def get_cell_value(sheet,cell_range):
    try:
        cell_value=sheet[cell_range].value
        return cell_value
    except:
        error="unable to get the value from range"
        print(error)

def create_worksheet(workbook_instance,workbook_path,sheet_name):
    wb=workbook_instance
    workbook_path=workbook_path
    try:
        sheet=wb.create_sheet(title=sheet_name)
        wb.save()
        wb.close()
        return sheet
    except:
        error="unable to create the worksheet"
        print(error)

def load_list_from_excel(file_path,sheet_name,number_of_columns,header):
    file_path, sheet_name, number_of_columns, header=file_path,sheet_name,number_of_columns,header
    list=""
    try:
        list=pd.read_excel(io=file_path,sheetname=sheet_name,header=header,parse_cols=number_of_columns)
        return list
    except:
        error="Unable to load the data int excel with the error: ",list
        return error

def add_list_into_excel(list,file_path,sheet_name):
    list, file_path, sheet_name=list,file_path,sheet_name
    try:
        success=list.to_excel(excel_writer=file_path,sheet_name=sheet_name)
        return success
    except:
        error="Unable to add the list into the given excel with the error: ",success
        return error

def load_list_from_csv_or_txt(file_path,header):
    file_path,header=file_path,header
    list=""
    try:
        list=pd.read_csv(file_path)
        return list
    except:
        error="Unable to load the data from csv/txt with the error: ",list
        return error

#debug code
def load_workbook(workbook_path):
    try:
        wb = openpyxl.load_workbook(workbook_path)
        sheet=wb.create_sheet(title="2910201999")
        wb.save(filename=workbook_path)
        return sheet
    except:
        error="unable to load excel workbook"
        print(error)


#print(load_workbook(r"C:\Users\43799850\Documents\Book3.xlsx"))

date_time_now=str(datetime.datetime.now())


print(date_time_now[0:4]+date_time_now[5:7]+date_time_now[8:10]+date_time_now[11:13]+date_time_now[14:16]+date_time_now[17:19])
print(date_time_now)